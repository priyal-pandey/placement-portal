from flask import Blueprint, request, jsonify
from models import Company, Drive, Application
from extensions import db
from openpyxl import Workbook
from io import BytesIO
from flask import send_file
import zipfile
import os
from io import BytesIO

company_bp = Blueprint("company", __name__)




# ---------------- DOWNLOAD ZIP ----------------
@company_bp.route("/download-resumes/<int:drive_id>", methods=["GET"])
def download_resumes(drive_id):
    apps = Application.query.filter_by(drive_id=drive_id).all()

    memory_file = BytesIO()

    with zipfile.ZipFile(memory_file, 'w') as zf:
        for a in apps:
            filename = a.resume or (a.student.resume if a.student else None)

            if filename:
                filepath = os.path.join("uploads", filename)

                if os.path.exists(filepath):
                    zf.write(filepath, arcname=filename)

    memory_file.seek(0)

    return send_file(
        memory_file,
        as_attachment=True,
        download_name=f"drive_{drive_id}_resumes.zip",
        mimetype="application/zip"
    )
# ---------------- EXPORT EXCEL ----------------
@company_bp.route("/export-excel/<int:drive_id>", methods=["GET"])
def export_excel(drive_id):
    apps = Application.query.filter_by(drive_id=drive_id).all()

    wb = Workbook()

    # Group by status
    grouped = {}
    for a in apps:
        grouped.setdefault(a.status, []).append(a)

    for status, items in grouped.items():
        ws = wb.create_sheet(title=status)

        ws.append([
            "Name", "Email", "CGPA", "Skills",
            "Branch", "Resume Link"
        ])

        for a in items:
            s = a.student

            resume_url = ""
            if a.resume:
                resume_url = f"{request.host_url}uploads/{a.resume}"
            elif s.resume:
                resume_url = f"{request.host_url}uploads/{s.resume}"

            ws.append([
                s.name,
                s.email,
                s.cgpa,
                s.skills,
                s.branch,
                resume_url
            ])

    # remove default sheet
    if "Sheet" in wb.sheetnames:
        std = wb["Sheet"]
        wb.remove(std)

    file_stream = BytesIO()
    wb.save(file_stream)
    file_stream.seek(0)

    return send_file(
        file_stream,
        as_attachment=True,
        download_name=f"drive_{drive_id}_applicants.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
# ---------------- CREATE DRIVE ----------------
@company_bp.route("/create-drive", methods=["POST"])
def create_drive():
    data = request.json

    company = Company.query.get(data["company_id"])

    if not company or not company.approved or company.is_blacklisted:
        return {"message": "Unauthorized"}, 403

    drive = Drive(
        company_id=company.id,
        title=data["title"],
        description=data.get("description"),
        requirements=data.get("requirements"),
        eligibility=data.get("eligibility"),
        salary=data.get("salary"),
        location=data.get("location"),
        deadline=data.get("deadline"),
        status="Pending"
    )

    db.session.add(drive)
    db.session.commit()

    return {"message": "Drive created, waiting for approval"}


# ---------------- GET DRIVES ----------------
@company_bp.route("/drives/<int:company_id>", methods=["GET"])
def get_company_drives(company_id):
    drives = Drive.query.filter_by(company_id=company_id).all()

    return jsonify([
        {
            "id": d.id,
            "title": d.title,
            "description": d.description,
            "requirements": d.requirements,
            "eligibility": d.eligibility,
            "salary": d.salary,
            "location": d.location,
            "deadline": d.deadline,
            "status": d.status
        } for d in drives
    ])


# ---------------- GET APPLICATIONS ----------------
@company_bp.route("/applications/<int:drive_id>", methods=["GET"])
def get_applications(drive_id):
    apps = Application.query.filter_by(drive_id=drive_id).all()

    data = []
    for a in apps:
        s = a.student

        data.append({
    "application_id": a.id,
    "status": a.status,
    "name": s.name,
    "email": s.email,
    "cgpa": s.cgpa,
    "skills": s.skills,
    "branch": s.branch,
    "phone": s.phone,  
    "passing_year": s.passing_year, 
    "resume": a.resume or s.resume,
    "cover_letter": a.cover_letter,
    "expected_salary": a.expected_salary
})

    return jsonify(data)


# ---------------- UPDATE STATUS ----------------
@company_bp.route("/update-status/<int:app_id>", methods=["POST"])
def update_status(app_id):
    data = request.json

    app = Application.query.get(app_id)

    if not app:
        return {"message": "Not found"}, 404

    app.status = data["status"]
    db.session.commit()

    return {"message": "Updated"}




# ---------------- VIEW PROFILE ----------------
@company_bp.route("/profile/<int:id>", methods=["GET"])
def get_company_profile(id):
    company = Company.query.get(id)

    if not company:
        return {"message": "Not found"}, 404

    return jsonify({
        "name": company.name,
        "email": company.email,
        "description": company.description,
        "industry": company.industry,
        "location": company.location,
        "website": company.website,
        "approved": company.approved
    })